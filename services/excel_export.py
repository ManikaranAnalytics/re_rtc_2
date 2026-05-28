"""
Live formula-driven RE-RTC Dispatch Workbook
Edit RTC Commitment / WTG Count / Solar AC MW in Config → everything recalculates.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── palette ───────────────────────────────────────────────────────────────────
DARK="0D1426"; HDR="1E2A45"; WIND="00B4D8"; SOLAR="F59E0B"
PSPD="8B5CF6"; PSPC="EC4899"; NET="10B981"; WARN="EF4444"; CURT="334155"

def _f(h): return PatternFill("solid", fgColor=h)
def _fn(b=False,c="E2E8F0",s=10,i=False): return Font(bold=b,color=c,size=s,name="Calibri",italic=i)
def _bd():
    t=Side(style="thin",color="1E2A45")
    return Border(left=t,right=t,top=t,bottom=t)
def _al(h="right",w=False): return Alignment(horizontal=h,vertical="center",wrap_text=w)
def _s(c,bg=DARK,fg="E2E8F0",b=False,s=10,a="right",i=False,f=None):
    c.fill=_f(bg); c.font=_fn(b,fg,s,i); c.border=_bd()
    c.alignment=_al(a,a=="center")
    if f: c.number_format=f

# Config sheet named cell refs (cross-sheet formulas)
RTC  = "Config!$B$6"
WTG  = "Config!$B$8"
SOL  = "Config!$B$11"
LOSS = "Config!$B$18"    # round-trip loss %
COMP = "Config!$B$19"    # min compliance ratio
CS   = "Config!$B$20"    # curtailment start block
CE   = "Config!$B$21"    # curtailment end block
ISOC = "Config!$B$23"    # initial SoC (EOD SoC from previous day)
DS   = "'Dispatch Schedule'"
RAW  = "'Raw Data'"


def build_excel(forecast_df, block_results, summary, rtc_range,
                rtc_commitment, wtg_count, solar_ac_mw, date_str,
                initial_soc_mwh=0.0,
                curtailment_enabled=True,
                curtailment_start_block=37, curtailment_end_block=64,
                roundtrip_loss_pct=20.0, min_compliance_ratio=0.75) -> bytes:
    wb = Workbook()
    ws_cfg  = wb.active;  ws_cfg.title = "Config"
    ws_raw  = wb.create_sheet("Raw Data")
    ws_disp = wb.create_sheet("Dispatch Schedule")
    ws_sum  = wb.create_sheet("Summary")
    _cfg(ws_cfg, date_str, wtg_count, solar_ac_mw, rtc_commitment, rtc_range,
         curtailment_enabled, curtailment_start_block, curtailment_end_block,
         roundtrip_loss_pct, min_compliance_ratio, initial_soc_mwh)
    _raw(ws_raw, forecast_df, wtg_count, solar_ac_mw)
    _disp(ws_disp, curtailment_start_block, curtailment_end_block, curtailment_enabled)
    _summ(ws_sum)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()


def _cfg(ws, date_str, wtg_count, solar_ac_mw, rtc, rtc_range,
         curtailment_enabled, curtailment_start_block, curtailment_end_block,
         roundtrip_loss_pct, min_compliance_ratio, initial_soc_mwh=0.0):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 44

    def title(span, v, fg="A5B4FC", s=13):
        ws.merge_cells(span); c=ws[span.split(":")[0]]
        c.value=v; _s(c,bg=DARK,fg=fg,b=True,s=s,a="center")

    def sec(r, v):
        ws.merge_cells(f"A{r}:C{r}"); c=ws.cell(r,1,v)
        _s(c,bg="1A2744",fg="818CF8",b=True,s=10,a="left")
        ws.row_dimensions[r].height=18

    def row(r, label, val, note="", edit=False, fg="F8FAFC", fmt=None):
        a=ws.cell(r,1,label); _s(a,bg=HDR,fg="94A3B8",b=True,s=10,a="left")
        b=ws.cell(r,2,val)
        _s(b,bg="0A1020" if edit else DARK,fg=fg,b=True,s=12 if edit else 10,a="right",f=fmt)
        if edit:
            g=Side(style="medium",color="10B981")
            b.border=Border(left=g,right=g,top=g,bottom=g)
        if note:
            nc=ws.cell(r,3,note); _s(nc,bg=DARK,fg="64748B",i=True,s=9,a="left")
        ws.row_dimensions[r].height=20

    title("A1:C1","RE-RTC Dispatch Optimizer — Live Configuration",s=14)
    ws.row_dimensions[1].height=30
    title("A2:C2","Aditya Birla Renewables  |  Hindalco Mahan 100 MW RTC Captive PPA",fg="64748B",s=10)
    ws.row_dimensions[2].height=16

    sec(4,"  ▸  SIMULATION INPUTS  (edit green-bordered cells — all sheets recalculate)")
    row(5,  "Simulation Date",           date_str)
    row(6,  "RTC Commitment (MW)  ✎",    rtc,
        "← EDIT THIS — flat daily commitment target",edit=True,fg="34D399",fmt="0.0")
    row(7,  "Min Compliance Floor (MW)", f"={RTC}*{COMP}",
        f"Auto: {int(min_compliance_ratio*100)}% of RTC — regulatory minimum")
    row(8,  "Wind Turbines (WTGs)  ✎",   wtg_count,
        "← EDIT THIS — active WTG count (1–59)",edit=True,fg="00D2FF",fmt="0")
    row(9,  "WTG Unit Capacity (MW)",    3.15,"Siemens Gamesa SG 3.15-114 (fixed)")
    row(10, "Total Wind Capacity (MW)",  f"={WTG}*3.15","Auto from WTG count")
    row(11, "Solar AC Capacity (MW)  ✎", solar_ac_mw,
        "← EDIT THIS — AC-side net capacity (5–175 MW)",edit=True,fg="F59E0B",fmt="0.0")

    sec(13,"  ▸  PSP STORAGE & DISPATCH CONSTANTS")
    row(14, "PSP Location",              "Orvakallu PSP, Andhra Pradesh")
    row(15, "Max Storage (MWh)",         360,"Hard ceiling (fixed)")
    row(16, "Max Charge Rate (MW)",      60, "Max draw from grid (fixed)")
    row(17, "Max Discharge Rate (MW)",   50, "Max injection to grid (fixed)")
    row(18, "Round-Trip Loss (%)",       roundtrip_loss_pct,
        "Total energy loss charging + discharging",fmt="0.0")
    row(19, "Min Compliance Ratio",      min_compliance_ratio,
        f"{int(min_compliance_ratio*100)}% of RTC is the regulatory floor",fmt="0.00")
    row(20, "Curtailment Start Block",   curtailment_start_block if curtailment_enabled else "DISABLED",
        "First block where Wind+Solar are zeroed")
    row(21, "Curtailment End Block",     curtailment_end_block if curtailment_enabled else "DISABLED",
        "Last block where Wind+Solar are zeroed")
    row(22, "Max Daily Cycles",          2.0,"CERC regulatory limit")

    sec(23,"  ▸  CARRY-FORWARD FROM PREVIOUS DAY")
    row(24, "Initial SoC (MWh)  ✎",     initial_soc_mwh,
        "← EOD SoC from previous day — sets Block 1 SoC Start in Dispatch Schedule",
        edit=True, fg="A78BFA", fmt="0.0")

    if rtc_range and "min_rtc_mw" in rtc_range:
        sec(26,"  ▸  MANIKARAN'S SUGGESTION  (dispatch-validated commitment analysis)")
        st=rtc_range.get("generation_stats",{})
        row(27,"Min Safe Commit (MW)",       rtc_range["min_rtc_mw"],
            f"{int(min_compliance_ratio*100)}% of P10 non-curtail gen",fg=WARN,fmt="0.00")
        row(28,"★ Recommended Commit (MW)",  rtc_range["recommended_rtc_mw"],
            "Max RTC → zero shortfall across all 96 blocks",fg="34D399",fmt="0.00")
        row(29,"Max Aggressive (MW)",        rtc_range["max_rtc_mw"],
            "P90 non-curtail gen (PSP backup for low blocks)",fg="818CF8",fmt="0.00")
        row(30,"Non-Curtail Gen P10 (MW)",   st.get("p10_mw",""),fmt="0.00")
        row(31,"Non-Curtail Gen Mean (MW)",  st.get("mean_mw",""),fmt="0.00")
        row(32,"Non-Curtail Gen P90 (MW)",   st.get("p90_mw",""),fmt="0.00")
        row(33,"Curtailment Loss (MWh/day)", rtc_range.get("curtailment_period_gen_lost_mwh",""),fmt="0.00")

    sec(35,"  ▸  HOW TO USE THIS WORKBOOK")
    ws.merge_cells("A36:C39")
    c=ws.cell(36,1)
    c.value=("1. Edit green/purple-bordered cells: RTC Commitment, WTG Count, Solar AC MW, Initial SoC.\n"
             "2. Go to 'Dispatch Schedule' — all 96 blocks recalculate automatically.\n"
             "3. Go to 'Summary' — all daily KPIs update from Dispatch Schedule.\n"
             "4. Do NOT edit 'Raw Data' — it holds meteorological data used by formulas.")
    _s(c,bg="0A1830",fg="94A3B8",i=True,s=10,a="left")
    c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
    ws.row_dimensions[36].height=70


def _raw(ws, forecast_df, wtg_count, solar_ac_mw):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    for col,w in {"A":7,"B":9,"C":20,"D":20,"E":14,"F":11}.items():
        ws.column_dimensions[col].width=w

    ws.merge_cells("A1:F1"); c=ws["A1"]
    c.value="Raw Data — Wind & Solar Source Data  (do NOT edit — auto-generated by backend)"
    _s(c,bg=DARK,fg="A5B4FC",b=True,s=11,a="center")
    ws.row_dimensions[1].height=22

    for col,lbl,clr in [("A","Block","94A3B8"),("B","Time","94A3B8"),
                         ("C","Per-WTG Power (kW)",WIND),
                         ("D","Solar Fraction\n(at any Solar AC MW)",SOLAR),
                         ("E","Wind Speed\nProjected (m/s)","7DD3FC"),
                         ("F","Curtailed?","64748B")]:
        c=ws[f"{col}2"]; c.value=lbl
        _s(c,bg=HDR,fg=clr,b=True,s=9,a="center")
    ws.row_dimensions[2].height=28

    for i,rd in enumerate(forecast_df.to_dict("records")):
        r=i+3
        is_c=bool(rd.get("curtail_flag",False))
        bg=CURT if is_c else ("111827" if i%2==0 else DARK)

        def rc(col,val,fmt=None,clr="CBD5E1",a="right"):
            c=ws[f"{col}{r}"]; c.value=val
            _s(c,bg=bg,fg=clr,s=9,a=a,f=fmt)

        wind_raw   = float(rd.get("wind_mw_raw",0))
        solar_raw  = float(rd.get("solar_mw_raw",0))
        per_wtg_kw = (wind_raw/wtg_count*1000.0) if wtg_count>0 else 0.0
        sol_frac   = (solar_raw/solar_ac_mw)      if solar_ac_mw>0 else 0.0

        rc("A",int(rd["block"]),              clr="94A3B8",a="center")
        rc("B",str(rd.get("time",""))[:5],    clr="94A3B8",a="center")
        rc("C",round(per_wtg_kw,4),           fmt="0.0000",clr=WIND)
        rc("D",round(sol_frac,6),             fmt="0.000000",clr=SOLAR)
        rc("E",round(float(rd.get("wind_speed",0)),2),fmt="0.00",clr="7DD3FC")
        rc("F","YES" if is_c else "NO",
           clr="F59E0B" if is_c else NET,a="center")


def _disp(ws, curtailment_start_block=37, curtailment_end_block=64, curtailment_enabled=True):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    COLS=[("A","Block",6,"94A3B8"),("B","Time",9,"94A3B8"),
          ("C","Curtailed?",11,"64748B"),
          ("D","Wind MW\n(live formula)",12,WIND),
          ("E","Solar MW\n(live formula)",13,SOLAR),
          ("F","Combined\nGen MW",12,"22D3EE"),
          ("G","Min Floor\nMW (75%)",11,WARN),
          ("H","SoC Start\nMWh",11,"6366F1"),
          ("I","PSP\nDischarge MW",13,PSPD),
          ("J","PSP\nCharge MW",12,PSPC),
          ("K","SoC End\nMWh",11,"6366F1"),
          ("L","Net Schedule\nMW",13,NET),
          ("M","RTC Target\nMW",11,WARN),
          ("N","RTM Surplus\nMW",12,"6B7280"),
          ("O","Compliant?",10,"334155")]

    ws.merge_cells("A1:O1"); c=ws["A1"]
    c.value=("Dispatch Schedule — 96 Time Blocks  |  All columns are live Excel formulas"
             "  |  Change Config sheet inputs to recalculate")
    _s(c,bg=DARK,fg="A5B4FC",b=True,s=12,a="center")
    ws.row_dimensions[1].height=24

    for col,lbl,w,clr in COLS:
        ws.column_dimensions[col].width=w
        c=ws[f"{col}2"]; c.value=lbl
        _s(c,bg=HDR,fg=clr,b=True,s=9,a="center")
    ws.row_dimensions[2].height=32

    # Build curtailment check formula using Config sheet values
    if curtailment_enabled:
        curt_check = f"AND(A{{r}}>={curtailment_start_block},A{{r}}<={curtailment_end_block})"
    else:
        curt_check = "FALSE"

    for i in range(96):
        r=i+3; pr=r-1
        is_c = curtailment_enabled and (curtailment_start_block <= i+1 <= curtailment_end_block)
        bg=CURT if is_c else ("111827" if i%2==0 else DARK)

        def fc(col,val,fmt=None,fg="CBD5E1",a="right",bold=False):
            c=ws[f"{col}{r}"]; c.value=val
            _s(c,bg=bg,fg=fg,s=9,a=a,f=fmt,b=bold)

        cf = curt_check.format(r=r)
        fc("A",f"={RAW}!A{r}",                a="center",fg="94A3B8")
        fc("B",f"={RAW}!B{r}",                a="center",fg="94A3B8")
        fc("C",f'=IF({cf},"YES","NO")',
           fg="F59E0B" if is_c else NET,a="center",bold=True)
        fc("D",f'=IF(C{r}="YES",0,{RAW}!C{r}/1000*{WTG})',
           fmt="0.00",fg=WIND)
        fc("E",f'=IF(C{r}="YES",0,{RAW}!D{r}*{SOL})',
           fmt="0.00",fg=SOLAR)
        fc("F",f"=D{r}+E{r}",                 fmt="0.00",fg="22D3EE",bold=True)
        # Min floor = COMP ratio * RTC
        fc("G",f"={RTC}*{COMP}",              fmt="0.00",fg=WARN)
        # Block 1: SoC start = Initial SoC from Config (carry-forward from previous day)
        # All other blocks: SoC start = SoC end of previous block
        soc_s = f"={ISOC}" if r == 3 else f"K{pr}"
        fc("H",f"={soc_s}",                   fmt="0.0", fg="818CF8")
        # Discharge loss factor = 1/(1 - LOSS/100)
        dlf = f"1/(1-{LOSS}/100)"
        fc("I",f"=IF(F{r}<G{r},MIN(G{r}-F{r},50,H{r}/(0.25*{dlf})),0)",
           fmt="0.00",fg=PSPD)
        fc("J",(f"=IF(F{r}>{RTC},"
                f"MIN(F{r}-{RTC},60,"
                f"(360-MAX(0,H{r}-I{r}*0.25*{dlf}))/0.25),0)"),
           fmt="0.00",fg=PSPC)
        fc("K",f"=MIN(360,MAX(0,H{r}-I{r}*0.25*{dlf})+J{r}*0.25)",
           fmt="0.0", fg="818CF8")
        fc("L",f"=F{r}+I{r}-J{r}",            fmt="0.00",fg=NET,bold=True)
        fc("M",f"={RTC}",                      fmt="0.00",fg=WARN)
        fc("N",f"=MAX(0,F{r}-{RTC}-J{r})",    fmt="0.00",fg="6B7280")
        fc("O",f'=IF(L{r}>={RTC}*{COMP}-0.0001,"✓ YES","✗ NO")',
           fg=NET,a="center",bold=True)

    # Totals row 99
    ws.merge_cells("A99:B99"); c=ws.cell(99,1,"TOTALS / AVERAGES")
    _s(c,bg=HDR,fg="A5B4FC",b=True,s=9,a="center")
    for col,formula,fmt in [
        ("D","=AVERAGE(D3:D98)","0.00"),
        ("E","=AVERAGE(E3:E98)","0.00"),
        ("F","=AVERAGE(F3:F98)","0.00"),
        ("I","=SUM(I3:I98)*0.25","0.00"),
        ("J","=SUM(J3:J98)*0.25","0.00"),
        ("K","=K98","0.0"),
        ("L","=AVERAGE(L3:L98)","0.00"),
        ("N","=SUM(N3:N98)*0.25","0.00"),
        ("O",'=COUNTIF(O3:O98,"✓ YES")&" / 96"',None),
    ]:
        c=ws[f"{col}99"]; c.value=formula
        _s(c,bg=HDR,fg="F8FAFC",b=True,s=9,a="right",f=fmt)
    ws.row_dimensions[99].height=18

    notes=[
        "FORMULA NOTES — how each column is calculated:",
        "D (Wind MW):      =IF(Curtailed,0, RawData.PerWTG_kW/1000 × Config.WTG_Count)",
        "E (Solar MW):     =IF(Curtailed,0, RawData.SolarFrac × Config.Solar_AC_MW)",
        "F (Combined Gen): =Wind_MW + Solar_MW",
        f"G (Min Floor):    =Config.RTC × Config.MinComplianceRatio  (75% regulatory threshold)",
        "H (SoC Start):    =0 for Block 1, then =SoC_End of previous block (chained)",
        "I (PSP Disch.):   =IF(Gen<Floor, MIN(Floor-Gen, 50MW, SoC/0.25/LossFactor), 0)",
        "J (PSP Charge):   =IF(Gen>RTC,   MIN(Gen-RTC, 60MW, space_in_tank), 0)",
        "K (SoC End):      =MIN(360, SoC_after_discharge + Charge_added)",
        "L (Net Schedule): =Gen_MW + PSP_Discharge - PSP_Charge",
        "N (RTM Surplus):  =MAX(0, Gen_MW - RTC_Target - PSP_Charge)",
        "O (Compliant?):   =IF(Net_Schedule >= RTC*MinComplianceRatio, YES, NO)",
    ]
    for j,note in enumerate(notes):
        c=ws.cell(101+j,1,note)
        _s(c,bg=DARK,fg="A5B4FC" if j==0 else "64748B",b=(j==0),i=(j>0),s=9,a="left")
    ws.merge_cells(f"A101:O{101+len(notes)-1}")


def _summ(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width=36
    ws.column_dimensions["B"].width=24
    ws.column_dimensions["C"].width=36

    ws.merge_cells("A1:C1"); c=ws["A1"]
    c.value="Daily Dispatch Summary — All values auto-calculated from Dispatch Schedule"
    _s(c,bg=DARK,fg="A5B4FC",b=True,s=13,a="center")
    ws.row_dimensions[1].height=28

    def sec(r,v):
        ws.merge_cells(f"A{r}:C{r}"); c=ws.cell(r,1,v)
        _s(c,bg="1A2744",fg="818CF8",b=True,s=10,a="left")
        ws.row_dimensions[r].height=18

    def row(r,label,formula,note="",fg="F8FAFC",fmt=None):
        a=ws.cell(r,1,label); _s(a,bg=HDR,fg="94A3B8",b=True,s=10,a="left")
        b=ws.cell(r,2,formula); _s(b,bg=DARK,fg=fg,b=True,s=11,a="right",f=fmt)
        if note:
            nc=ws.cell(r,3,note); _s(nc,bg=DARK,fg="64748B",i=True,s=9,a="left")
        ws.row_dimensions[r].height=20

    sec(3,"  ▸  CONFIGURATION (from Config sheet)")
    row(4, "RTC Commitment (MW)",         f"={RTC}",         fmt="0.00",fg="34D399")
    row(5, "Min Compliance Floor (MW)",   f"={RTC}*{COMP}",  fmt="0.00")
    row(6, "Min Compliance Ratio",        f"={COMP}",        fmt="0.0%")
    row(7, "WTG Count",                   f"={WTG}",         fmt="0")
    row(8, "Solar AC Capacity (MW)",      f"={SOL}",         fmt="0.0",fg=SOLAR)
    row(9, "Round-Trip Loss (%)",         f"={LOSS}",        fmt="0.0")

    sec(11,"  ▸  GENERATION TOTALS")
    row(12,"Total Wind Generation (MWh)", f"=SUM({DS}!D3:D98)*0.25",fmt="0.00",fg=WIND)
    row(13,"Total Solar Generation (MWh)",f"=SUM({DS}!E3:E98)*0.25",fmt="0.00",fg=SOLAR)
    row(14,"Total Combined Gen (MWh)",    f"=SUM({DS}!F3:F98)*0.25",fmt="0.00",fg="22D3EE")
    row(15,"Avg Gen per Block (MW)",      f"=AVERAGE({DS}!F3:F98)", fmt="0.00")

    sec(17,"  ▸  PSP STORAGE DISPATCH")
    row(18,"Total PSP Discharged (MWh)",  f"=SUM({DS}!I3:I98)*0.25",fmt="0.00",fg=PSPD)
    row(19,"Total PSP Charged (MWh)",     f"=SUM({DS}!J3:J98)*0.25",fmt="0.00",fg=PSPC)
    row(20,"PSP Usable Energy (MWh)",     f"=SUM({DS}!J3:J98)*0.25*(1-{LOSS}/100)",
        "Actual recoverable energy after round-trip losses",fmt="0.00",fg=PSPC)
    row(21,"PSP Cycles Used",             f"=SUM({DS}!J3:J98)*0.25/360",fmt="0.00")
    row(22,"End-of-Day SoC (MWh)",        f"={DS}!K98",              fmt="0.0",fg="818CF8")
    row(23,"End-of-Day SoC (%)",          f"={DS}!K98/360",          fmt="0.0%",fg="818CF8")

    sec(25,"  ▸  COMPLIANCE & DELIVERY")
    row(26,"Compliant Blocks",            f'=COUNTIF({DS}!O3:O98,"✓ YES")&" / 96"',fg=NET)
    row(27,"Compliance Rate (%)",         f'=COUNTIF({DS}!O3:O98,"✓ YES")/96',fmt="0.0%",fg=NET)
    row(28,"Total Net Schedule (MWh)",    f"=SUM({DS}!L3:L98)*0.25",fmt="0.00",fg=NET)
    row(29,"Total RTM Surplus (MWh)",     f"=SUM({DS}!N3:N98)*0.25",fmt="0.00",fg="6B7280",
        note="Exportable generation above RTC target")
    row(30,"Fully Compliant Day?",
        f'=IF(COUNTIF({DS}!O3:O98,"✓ YES")=96,"✓  YES — 100% blocks met","✗  NO — shortfall blocks exist")',
        fg=NET)

    sec(32,"  ▸  CARRY-FORWARD (PSP SoC Roll)")
    row(33,"Initial SoC (MWh)",           f"={ISOC}",
        "EOD SoC from previous day — carry budget into today",fmt="0.0",fg="A78BFA")
    row(34,"Carry Budget Discharged (MWh)",
        f"=MAX(0,{ISOC}-{DS}!K3+{DS}!I3*0.25*(1/(1-{LOSS}/100)))",
        "Carry energy consumed in Block 1 (approx)",fmt="0.0",fg="A78BFA")
    row(35,"End-of-Day SoC → next day carry",f"={DS}!K98",
        "Pass this value as Initial SoC for next day's simulation",fmt="0.0",fg="A78BFA")
